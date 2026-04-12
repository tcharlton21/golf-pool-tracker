"""
Tests for Excel parser using in-memory workbooks (no disk files required).
"""

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from services.excel_parser import (
    PIPER_GROUP_LABELS,
    _parse_piper_form,
    _parse_piper_matrix,
    parse_marshalek,
)


def _make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """Build an in-memory xlsx workbook from a dict of {sheet_name: rows}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseMarshalek:
    def test_basic_parse(self):
        xlsx = _make_xlsx({
            "Picks": [
                ["Entrant", "Pick1", "Pick2", "Pick3", "Pick4", "Pick5"],
                ["Alan Superfine", "Dechambeau, Bryson", "Fleetwood, Tommy", "Rose, Justin", "Aberg, Ludvig", "Lee, Min Woo"],
                ["Bob Smith", "Scheffler, Scottie", "McIlroy, Rory", None, None, None],
            ]
        })
        result = parse_marshalek(xlsx)
        assert len(result.entrants) == 2
        alan = result.entrants[0]
        assert alan.name == "Alan Superfine"
        assert len(alan.picks) == 5
        assert alan.picks[0].golfer_name == "Bryson Dechambeau"
        assert alan.picks[0].golfer_name_raw == "Dechambeau, Bryson"
        assert alan.picks[0].group_label is None
        assert alan.picks[0].pick_order == 1

    def test_skips_empty_rows(self):
        xlsx = _make_xlsx({
            "Picks": [
                ["Entrant", "Pick1", "Pick2"],
                ["Alice", "Scheffler, Scottie", "McIlroy, Rory"],
                [None, None, None],
                [" ", None, None],
                ["Bob", "Rahm, Jon", None],
            ]
        })
        result = parse_marshalek(xlsx)
        assert len(result.entrants) == 2

    def test_partial_picks(self):
        xlsx = _make_xlsx({
            "Picks": [
                ["Entrant", "P1", "P2", "P3"],
                ["Alice", "Scheffler, Scottie", None, None],
            ]
        })
        result = parse_marshalek(xlsx)
        assert len(result.entrants[0].picks) == 1

    def test_missing_sheet_raises(self):
        xlsx = _make_xlsx({"WrongSheet": [["A", "B"]]})
        with pytest.raises(ValueError, match="Could not find 'Picks' sheet"):
            parse_marshalek(xlsx)

    def test_no_entrants_raises(self):
        xlsx = _make_xlsx({"Picks": [["Header"]]})
        with pytest.raises(ValueError, match="No entrants parsed"):
            parse_marshalek(xlsx)


class TestParsePiperMatrix:
    def _make_matrix_ws(self, entrant_names: list[str], golfer_to_picks: dict[str, list[str]]):
        """Build a worksheet in Player Selection Sheet matrix format."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Player Selection Sheet"

        # Row 1: blank, then entrant names
        header = [None] + entrant_names
        ws.append(header)

        # Rows 2+: golfer name in col A, "x" in entrant columns
        all_golfers = list(golfer_to_picks.keys())
        for golfer in all_golfers:
            row = [golfer]
            for entrant in entrant_names:
                row.append("x" if entrant in golfer_to_picks[golfer] else None)
            ws.append(row)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_basic_matrix(self):
        xlsx = self._make_matrix_ws(
            entrant_names=["Alice", "Bob"],
            golfer_to_picks={
                "Scottie Scheffler": ["Alice", "Bob"],
                "Rory McIlroy": ["Alice"],
                "Jon Rahm": ["Bob"],
            },
        )
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb.active
        result = _parse_piper_matrix(ws)

        assert len(result.entrants) == 2
        alice = next(e for e in result.entrants if e.name == "Alice")
        assert len(alice.picks) == 2
        golfer_names = [p.golfer_name for p in alice.picks]
        assert "Scottie Scheffler" in golfer_names
        assert "Rory McIlroy" in golfer_names

    def test_group_labels_assigned(self):
        xlsx = self._make_matrix_ws(
            entrant_names=["Alice"],
            golfer_to_picks={
                f"Player {i}": ["Alice"] for i in range(10)
            },
        )
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb.active
        result = _parse_piper_matrix(ws)

        alice = result.entrants[0]
        assert len(alice.picks) == 10
        labels = [p.group_label for p in alice.picks]
        assert labels == PIPER_GROUP_LABELS


class TestParsePiperForm:
    def _make_form_ws(self, rows: list[dict]):
        """Build a form response worksheet."""
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["Timestamp", "Email Address", "Name"] + PIPER_GROUP_LABELS
        ws.append(headers)

        for row in rows:
            data = [
                "2026-04-10T10:00:00",
                row.get("email", "test@test.com"),
                row["name"],
            ] + [row.get(label) for label in PIPER_GROUP_LABELS]
            ws.append(data)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_basic_form_parse(self):
        xlsx = self._make_form_ws([
            {
                "name": "Alice",
                "email": "alice@test.com",
                "A-1": "Scheffler, Scottie",
                "A-2": "McIlroy, Rory",
                "B-1": "Rahm, Jon",
                "B-2": "Fleetwood, Tommy",
                "C-1": "Aberg, Ludvig",
                "C-2": "Rose, Justin",
                "D-1": "Kim, Tom",
                "D-2": "Burns, Sam",
                "E-1": "Clark, Wyndham",
                "E-2": "Hoge, Tom",
            }
        ])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb.active
        result = _parse_piper_form(ws)

        assert len(result.entrants) == 1
        alice = result.entrants[0]
        assert alice.name == "Alice"
        assert len(alice.picks) == 10
        assert alice.picks[0].golfer_name == "Scottie Scheffler"
        assert alice.picks[0].group_label == "A-1"

    def test_empty_picks_skipped(self):
        xlsx = self._make_form_ws([
            {"name": "Incomplete", "A-1": "Scheffler, Scottie"},
            {"name": "NoPicksAtAll"},
        ])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb.active
        result = _parse_piper_form(ws)

        # NoPicksAtAll has no picks → skipped
        names = [e.name for e in result.entrants]
        assert "NoPicksAtAll" not in names
        assert "Incomplete" in names


class TestNormalizeExcelName:
    def test_last_comma_first(self):
        from services.excel_parser import _find_sheet
        from services.player_matcher import normalize_excel_name
        assert normalize_excel_name("Scheffler, Scottie") == "Scottie Scheffler"

    def test_already_normalized(self):
        from services.player_matcher import normalize_excel_name
        assert normalize_excel_name("Scottie Scheffler") == "Scottie Scheffler"

    def test_mcilroy(self):
        from services.player_matcher import normalize_excel_name
        assert normalize_excel_name("McIlroy, Rory") == "Rory McIlroy"

    def test_strips_whitespace(self):
        from services.player_matcher import normalize_excel_name
        assert normalize_excel_name("  Rose, Justin  ") == "Justin Rose"
