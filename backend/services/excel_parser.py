"""
Parse Excel pick sheets for Marshalek and Piper pool formats.

Marshalek format ("Picks" sheet):
  Row 1: headers (skipped)
  Col A: entrant name
  Cols B-F: picks in "Last, First" format (col F often empty)

Piper format — two possible sub-formats:
  1. "Player Selection Sheet" — matrix with player names in rows,
     entrant names in columns, "x" marking a pick
  2. "Sheet1" — raw Google Form responses with columns labeled
     "A-1", "A-2", "B-1", "B-2", "C-1", "C-2", "D-1", "D-2", "E-1", "E-2"
"""

import io
import logging
from dataclasses import dataclass, field

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from services.player_matcher import normalize_excel_name

logger = logging.getLogger(__name__)


@dataclass
class PickPreview:
    golfer_name: str       # normalized: "Scottie Scheffler"
    golfer_name_raw: str   # original cell value
    group_label: str | None
    pick_order: int


@dataclass
class EntrantPreview:
    name: str
    picks: list[PickPreview] = field(default_factory=list)


@dataclass
class ParseResult:
    entrants: list[EntrantPreview]
    warnings: list[str]


# Piper group labels in order
PIPER_GROUP_LABELS = ["A-1", "A-2", "B-1", "B-2", "C-1", "C-2", "D-1", "D-2", "E-1", "E-2"]


def parse_marshalek(file_bytes: bytes) -> ParseResult:
    """
    Parse the Marshalek pool Excel file.
    Expected sheet: "Picks"
    Structure: col A = entrant name, cols B-F = picks ("Last, First")
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    sheet_name = _find_sheet(wb, ["Picks", "picks", "Sheet1"])
    if sheet_name is None:
        raise ValueError(f"Could not find 'Picks' sheet. Available: {wb.sheetnames}")

    ws: Worksheet = wb[sheet_name]
    warnings: list[str] = []
    entrants: list[EntrantPreview] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        entrant_name = row[0]
        if not entrant_name or not str(entrant_name).strip():
            continue

        entrant_name = str(entrant_name).strip()
        picks: list[PickPreview] = []

        pick_cells = row[1:6]  # cols B-F (indices 1-5)
        for pick_order, cell_val in enumerate(pick_cells, start=1):
            if not cell_val or not str(cell_val).strip():
                continue
            raw = str(cell_val).strip()
            try:
                normalized = normalize_excel_name(raw)
            except Exception:
                warnings.append(f"Row {row_idx}: could not normalize pick '{raw}' for {entrant_name}")
                normalized = raw
            picks.append(PickPreview(
                golfer_name=normalized,
                golfer_name_raw=raw,
                group_label=None,
                pick_order=pick_order,
            ))

        if len(picks) == 0:
            warnings.append(f"Row {row_idx}: entrant '{entrant_name}' has no picks — skipped")
            continue

        entrants.append(EntrantPreview(name=entrant_name, picks=picks))

    if not entrants:
        raise ValueError("No entrants parsed from Marshalek file — check sheet format")

    return ParseResult(entrants=entrants, warnings=warnings)


def parse_piper(file_bytes: bytes) -> ParseResult:
    """
    Parse the Piper pool Excel file.
    Tries "Player Selection Sheet" (matrix) first, then "Sheet1" (form responses).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    matrix_sheet = _find_sheet(wb, ["Player Selection Sheet", "Player Selections", "Entry Sheet", "Selections", "Matrix"])
    form_sheet = _find_sheet(wb, ["Sheet1", "Form Responses", "Responses"])

    if matrix_sheet:
        result = _parse_piper_matrix(wb[matrix_sheet])
        if result.entrants:
            return result

    if form_sheet:
        result = _parse_piper_form(wb[form_sheet])
        if result.entrants:
            return result

    raise ValueError(
        f"Could not parse Piper file. Available sheets: {wb.sheetnames}. "
        "Expected 'Player Selection Sheet' (matrix) or 'Sheet1' (form responses)."
    )


def _parse_piper_matrix(ws: Worksheet) -> ParseResult:
    """
    Parse the "Player Selection Sheet" matrix.
    Row 1: entrant names in columns (starting from some column)
    Subsequent rows: player names in col A or B, "x" marks in entrant columns
    """
    warnings: list[str] = []
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return ParseResult(entrants=[], warnings=["Empty sheet"])

    # Find the header row: the row where the most non-empty string cells appear
    # (entrant names are in row 1, player names start row 2+)
    header_row = all_rows[0]

    # Find the first column with an entrant name (skip leading meta columns).
    # Row 2 holds each entrant's pick count (typically 10 in Piper) — use it as
    # the authoritative signal: a real entrant column has a numeric count in row 2.
    count_row = all_rows[1] if len(all_rows) > 1 else ()
    entrant_col_start = None
    for col_idx, cell in enumerate(header_row):
        if not cell or not str(cell).strip():
            continue
        val = str(cell).strip()
        if val in ("Player", "Name") or val.lower().startswith(("total", "group")):
            continue
        if len(val) <= 2:
            continue
        count_cell = count_row[col_idx] if col_idx < len(count_row) else None
        if not isinstance(count_cell, (int, float)) or count_cell <= 0:
            continue
        entrant_col_start = col_idx
        break

    if entrant_col_start is None:
        return ParseResult(entrants=[], warnings=["Could not find entrant columns in matrix sheet"])

    # Collect entrant names and their column indices
    entrant_cols: list[tuple[int, str]] = []
    for col_idx in range(entrant_col_start, len(header_row)):
        cell = header_row[col_idx]
        if cell and str(cell).strip():
            entrant_cols.append((col_idx, str(cell).strip()))

    if not entrant_cols:
        return ParseResult(entrants=[], warnings=["No entrant columns found in header row"])

    # Build entrant pick lists — we need to find golfer names
    # The golfer names may be in column 0 or 1; detect which has more golf player names
    picks_by_entrant: dict[str, list[str]] = {name: [] for _, name in entrant_cols}

    for row in all_rows[1:]:
        # Find golfer name (usually in the first non-empty cell of the row, col 0 or 1)
        golfer_raw: str | None = None
        for cell in row[:entrant_col_start]:
            if cell and str(cell).strip():
                golfer_raw = str(cell).strip()
                break

        if not golfer_raw:
            continue

        # Check which entrant columns have an "x" (case-insensitive)
        for col_idx, entrant_name in entrant_cols:
            if col_idx < len(row):
                cell_val = row[col_idx]
                if cell_val and str(cell_val).strip().lower() in ("x", "✓", "yes", "1"):
                    picks_by_entrant[entrant_name].append(golfer_raw)

    entrants: list[EntrantPreview] = []
    for col_idx, entrant_name in entrant_cols:
        raw_picks = picks_by_entrant[entrant_name]
        if len(raw_picks) == 0:
            warnings.append(f"Entrant '{entrant_name}' has no picks in matrix — skipped")
            continue

        picks: list[PickPreview] = []
        for order, raw in enumerate(raw_picks, start=1):
            normalized = normalize_excel_name(raw)
            group_label = PIPER_GROUP_LABELS[order - 1] if order <= len(PIPER_GROUP_LABELS) else None
            picks.append(PickPreview(
                golfer_name=normalized,
                golfer_name_raw=raw,
                group_label=group_label,
                pick_order=order,
            ))
        entrants.append(EntrantPreview(name=entrant_name, picks=picks))

    return ParseResult(entrants=entrants, warnings=warnings)


def _parse_piper_form(ws: Worksheet) -> ParseResult:
    """
    Parse "Sheet1" which contains raw Google Form responses.
    Row 1: column headers (Timestamp, Email, Name, A-1, A-2, B-1, B-2, ...)
    Rows 2+: one form submission per row
    """
    warnings: list[str] = []
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < 2:
        return ParseResult(entrants=[], warnings=["Form sheet has no data rows"])

    header = [str(cell).strip() if cell else "" for cell in all_rows[0]]

    # Find the column index for "Name" or "Your Name"
    name_col = None
    for idx, h in enumerate(header):
        if h.lower() in ("name", "your name", "full name", "entrant"):
            name_col = idx
            break

    if name_col is None:
        # Fall back: assume col 2 is name (after Timestamp, Email)
        name_col = 2
        warnings.append("Could not find 'Name' column header — assuming column 3 (index 2)")

    # Find pick group columns (A-1, A-2, B-1, B-2, C-1, C-2, D-1, D-2, E-1, E-2)
    group_cols: dict[str, int] = {}
    for idx, h in enumerate(header):
        clean = h.strip().upper()
        # Match patterns like "A-1", "Selection A-1", "[A-1]", "A1"
        for label in PIPER_GROUP_LABELS:
            label_variants = [label, label.replace("-", ""), f"[{label}]", f"Selection {label}"]
            if clean in [v.upper() for v in label_variants] or label.upper() in clean:
                group_cols[label] = idx
                break

    if len(group_cols) < 10:
        warnings.append(
            f"Only found {len(group_cols)} group columns in form sheet (expected 10). "
            f"Headers: {header}"
        )

    entrants: list[EntrantPreview] = []
    for row_idx, row in enumerate(all_rows[1:], start=2):
        if not row or not any(row):
            continue

        raw_name = row[name_col] if name_col < len(row) else None
        if not raw_name or not str(raw_name).strip():
            continue
        entrant_name = str(raw_name).strip()

        picks: list[PickPreview] = []
        for pick_order, label in enumerate(PIPER_GROUP_LABELS, start=1):
            col_idx = group_cols.get(label)
            if col_idx is None or col_idx >= len(row):
                continue
            cell_val = row[col_idx]
            if not cell_val or not str(cell_val).strip():
                continue
            raw = str(cell_val).strip()
            normalized = normalize_excel_name(raw)
            picks.append(PickPreview(
                golfer_name=normalized,
                golfer_name_raw=raw,
                group_label=label,
                pick_order=pick_order,
            ))

        if len(picks) == 0:
            warnings.append(f"Row {row_idx}: entrant '{entrant_name}' has no picks — skipped")
            continue

        entrants.append(EntrantPreview(name=entrant_name, picks=picks))

    return ParseResult(entrants=entrants, warnings=warnings)


def _find_sheet(wb: openpyxl.Workbook, candidates: list[str]) -> str | None:
    """Return the first sheet name from candidates that exists in the workbook."""
    sheet_names_lower = {name.lower(): name for name in wb.sheetnames}
    for candidate in candidates:
        if candidate.lower() in sheet_names_lower:
            return sheet_names_lower[candidate.lower()]
    return None
